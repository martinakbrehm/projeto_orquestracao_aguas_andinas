#!/usr/bin/env python3
import os
import re
import csv
import argparse
from pathlib import Path


def find_result_files(folder: Path):
    files = list(folder.glob('*.xlsx'))
    files += list(folder.glob('*.csv'))
    return sorted(files)


def is_row_nonempty(row):
    return any((cell is not None and str(cell).strip() != '') for cell in row)


def row_has_phone(row, min_digits=8):
    # prefer explicit phone-like patterns (country code +56, leading 9 for mobiles)
    phone_re = re.compile(r"(?:(?:\+?56)[\s\-\.]*)?9[\s\-\.]*\d{2}[\s\-\.]*\d{3}[\s\-\.]*\d{3}")
    generic_re = re.compile(r"\d{8,12}")
    for cell in row:
        if cell is None:
            continue
        s = str(cell)
        if phone_re.search(s):
            return True
        # generic digits: accept only if starts with 9 or 56 (to reduce false positives)
        for m in generic_re.finditer(re.sub(r"\D", "", s)):
            digits = m.group(0)
            if len(digits) >= min_digits and (digits.startswith('9') or digits.startswith('56')):
                return True
    return False


def process_file(path: Path, min_digits=8):
    total_rows = 0
    rows_with_phone = 0

    if path.suffix.lower() == '.xlsx':
        try:
            from openpyxl import load_workbook
        except Exception:
            print("Erro: biblioteca 'openpyxl' não encontrada. Instale com: pip install openpyxl")
            raise

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        for sheet in wb.worksheets:
            iter_rows = sheet.iter_rows(values_only=True)
            try:
                first = next(iter_rows)
            except StopIteration:
                continue

            header = any(isinstance(c, str) for c in first if c is not None)
            # count first row if not empty and not header
            if is_row_nonempty(first) and not header:
                total_rows += 1
                if row_has_phone(first, min_digits=min_digits):
                    rows_with_phone += 1

            for row in iter_rows:
                if not is_row_nonempty(row):
                    continue
                total_rows += 1
                if row_has_phone(row, min_digits=min_digits):
                    rows_with_phone += 1

    elif path.suffix.lower() == '.csv':
        import csv as _csv
        with path.open(newline='', encoding='utf-8') as fh:
            reader = _csv.reader(fh)
            try:
                first = next(reader)
            except StopIteration:
                return 0, 0

            # detect header if any alphabetic char in first row
            header = any(re.search('[A-Za-z]', str(c)) for c in first if c is not None)
            if any(str(c).strip() != '' for c in first) and not header:
                total_rows += 1
                if row_has_phone(first, min_digits=min_digits):
                    rows_with_phone += 1

            for row in reader:
                if not any(str(c).strip() != '' for c in row):
                    continue
                total_rows += 1
                if row_has_phone(row, min_digits=min_digits):
                    rows_with_phone += 1

    else:
        # unsupported
        return 0, 0

    return total_rows, rows_with_phone


def main():
    parser = argparse.ArgumentParser(description='Contar linhas e telefones em arquivos .xlsx em dados/resultados')
    parser.add_argument('--folder', '-f', default='dados/resultados', help='Pasta com os arquivos .xlsx')
    parser.add_argument('--min-digits', '-m', type=int, default=8, help='Mínimo de dígitos para considerar como telefone')
    parser.add_argument('--report', '-r', default='dados/resultados/phone_report.csv', help='Caminho do relatório CSV de saída')
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists() or not folder.is_dir():
        print(f"Pasta não encontrada: {folder}")
        return

    files = find_result_files(folder)
    if not files:
        print(f"Nenhum arquivo .xlsx encontrado em {folder}")
        return

    summary = []
    grand_total = 0
    grand_with_phone = 0

    for f in files:
        try:
            total, with_phone = process_file(f, min_digits=args.min_digits)
        except Exception as e:
            print(f"Falha ao processar {f}: {e}")
            continue
        summary.append((str(f), total, with_phone))
        grand_total += total
        grand_with_phone += with_phone

    # write csv report
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open('w', newline='', encoding='utf-8') as csvf:
        writer = csv.writer(csvf)
        writer.writerow(['file', 'total_rows', 'rows_with_phone'])
        for row in summary:
            writer.writerow(row)
        writer.writerow(['TOTAL', grand_total, grand_with_phone])

    # print concise summary
    print('Resumo:')
    for path, total, with_phone in summary:
        pct = (with_phone / total * 100) if total else 0
        print(f"- {os.path.basename(path)}: {total} linhas, {with_phone} com telefone ({pct:.1f}%)")
    grand_pct = (grand_with_phone / grand_total * 100) if grand_total else 0
    print(f"Total geral: {grand_total} linhas; {grand_with_phone} com telefone ({grand_pct:.1f}%)")
    print(f"Relatório salvo em: {report_path}")


if __name__ == '__main__':
    main()
