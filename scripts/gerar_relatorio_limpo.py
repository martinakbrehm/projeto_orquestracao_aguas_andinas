"""
gerar_relatorio_limpo.py
=========================
Gera relatório Excel de clientes por cidade/região a partir do arquivo
ENTREGA BASE NOMBRE DIRECCION FECH NAC_LIMPO.txt (já limpo).

Aplica allowlist das 16 regiões oficiais do Chile.
Conta apenas RUTs únicos (cliente único).

Saída: dados/resultados/relatorio_clientes_por_cidade_LIMPO_YYYYMMDD_HHMMSS.xlsx
"""
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

LIMPO  = ROOT / "dados" / "bases" / "ENTREGA BASE NOMBRE DIRECCION FECH NAC_LIMPO5.txt"
OUTPUT = ROOT / "dados" / "resultados"
OUTPUT.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Regiões oficiais do Chile (16) — strings exatas como estão no arquivo limpo
# ---------------------------------------------------------------------------
REGIOES_VALIDAS = {
    "METROPOLITANA DE SANTIAGO",
    "DE VALPARAÍSO",
    "DEL BIOBÍO",
    "DE LA ARAUCANÍA",
    "DEL LIBERTADOR GRAL. BERNARDO O'HIGGINS",
    "DE LOS LAGOS",
    "DE COQUIMBO",
    "DE ANTOFAGASTA",
    "DE LOS RÍOS",
    "DE TARAPACÁ",
    "DE ARICA Y PARINACOTA",
    "DE MAGALLANES Y LA ANTÁRTICA CHILENA",
    "DE AYSÉN DEL GRAL. CARLOS IBÁÑEZ DEL CAMPO",
    "DEL MAULE",
    "DE ÑUBLE",
    "DE ATACAMA",
}

RE_DIGIT    = re.compile(r"[0-9]")
CHARS_INV   = set(".#-(/,")


def clean_comuna(v: str):
    v = (v or "").strip()
    if not v or len(v) < 3:
        return None
    if RE_DIGIT.search(v):
        return None
    if v[0] in CHARS_INV:
        return None
    return v


def main():
    print(f"Lendo: {LIMPO.name}")
    ruts_vistos: set = set()
    rows = []

    with open(LIMPO, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rut = (row.get("rut") or "").strip()
            if not rut or rut in ruts_vistos:
                continue
            ruts_vistos.add(rut)

            region = (row.get("region") or "").strip()
            region = region if region in REGIOES_VALIDAS else None

            rows.append({
                "region": region,
                "comuna": clean_comuna(row.get("comuna", "")),
            })

    df = pd.DataFrame(rows)
    print(f"Clientes únicos: {len(df):,}")

    # ── Por cidade ──────────────────────────────────────────────────────
    por_cidade = (
        df.groupby(["region", "comuna"], dropna=False)
        .size().reset_index(name="total")
        .fillna({"region": "(sem região)", "comuna": "(sem cidade)"})
        .sort_values(["region", "comuna"])
        .reset_index(drop=True)
    )
    por_cidade.columns = ["Região", "Cidade (Comuna)", "Total Clientes"]

    # ── Por região ───────────────────────────────────────────────────────
    por_regiao = (
        df.groupby("region", dropna=False)
        .size().reset_index(name="total")
        .fillna({"region": "(sem região)"})
        .sort_values("total", ascending=False)
        .reset_index(drop=True)
    )
    por_regiao.columns = ["Região", "Total Clientes"]

    # ── Resumo ───────────────────────────────────────────────────────────
    agora = datetime.now()
    df_resumo = pd.DataFrame([
        {"Informação": "Total clientes únicos",   "Valor": len(df)},
        {"Informação": "Cidades (comunas)",        "Valor": len(por_cidade)},
        {"Informação": "Regiões (inc. sem região)","Valor": len(por_regiao)},
        {"Informação": "Arquivo base",             "Valor": LIMPO.name},
        {"Informação": "Data de geração",          "Valor": agora.strftime("%d/%m/%Y %H:%M:%S")},
    ])

    # ── Excel ────────────────────────────────────────────────────────────
    ts  = agora.strftime("%Y%m%d_%H%M%S")
    out = OUTPUT / f"relatorio_clientes_por_cidade_LIMPO_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        por_cidade.to_excel(writer, sheet_name="Por Cidade", index=False)
        por_regiao.to_excel(writer, sheet_name="Por Região", index=False)
        df_resumo.to_excel(writer,  sheet_name="Resumo",     index=False)

        wb    = writer.book
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hal   = Alignment(horizontal="center", vertical="center")

        for sn, dt in [("Por Cidade", por_cidade), ("Por Região", por_regiao), ("Resumo", df_resumo)]:
            ws = wb[sn]
            for cell in ws[1]:
                cell.font = hfont; cell.fill = hfill; cell.alignment = hal
            for i, col in enumerate(dt.columns, 1):
                ml = max(len(str(col)),
                         dt.iloc[:, i - 1].astype(str).str.len().max() if len(dt) else 0)
                ws.column_dimensions[ws.cell(1, i).column_letter].width = min(ml + 4, 55)

    print(f"\nRelatório gerado: {out.name}")
    print(f"  Cidades  : {len(por_cidade):,}")
    print(f"  Regiões  : {len(por_regiao):,}")
    print()
    print("Distribuição por região:")
    print(por_regiao.to_string(index=False))


if __name__ == "__main__":
    main()
